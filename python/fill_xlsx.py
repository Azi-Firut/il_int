import json
import openpyxl
import os
import re
import sys
import win32com.client

def fill_excel(data):
    # Extract the Excel file path
    excel_path = data.pop('adressXLSX')
    pdf_path = re.sub(r'\.xlsx$', '.pdf', excel_path)

    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Helper function to get the top-left cell of a merged cell range
    def get_top_left_cell(cell):
        for merged_cell_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_cell_range:
                return merged_cell_range.coord.split(":")[0]
        return cell.coordinate

    # Fill the specified cells with the provided data
    for cell, value in data.items():
        top_left_cell = get_top_left_cell(ws[cell])
        ws[top_left_cell] = value

    # Save the modified Excel file
    modified_excel_path = re.sub(r'\.xlsx$', '.xlsx', excel_path)
    wb.save(modified_excel_path)

    # Convert the modified Excel file to PDF using win32com
    excel_to_pdf(modified_excel_path, pdf_path)

def excel_to_pdf(excel_path, pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(excel_path)

    try:
        wb.ExportAsFixedFormat(0, pdf_path)
    except Exception as e:
        print(f"Failed to convert to PDF: {e}")
    finally:
        wb.Close(False)
        excel.Quit()

# Check if the script is run with command-line arguments
if len(sys.argv) > 1:
    # The first argument is the script name, the second is the JSON data
    credentials_json = sys.argv[1]
    data = json.loads(credentials_json)

    fill_excel(data)
else:
    print("No data provided")
